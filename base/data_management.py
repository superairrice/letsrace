from django.db import connection
import openpyxl as op


def get_krafile(rcity, rdate1, rdate2, fcode, fstatus):

    try:
        cursor = connection.cursor()

        strSql = (
            """ 
            SELECT a.fname,   
                    @row:=@row+1 as No,
                " " as rcheck,
                    a.fpath,
                    a.rdate,   
                    a.fcode,   
                    a.fstatus,   
                    a.in_date  
            FROM krafile a,(select @row :=0 from dual) b
            WHERE Left( Right( fname , 6), 2) like '"""
            + rcity
            + """%'
            AND rdate between '"""
            + rdate1
            + """' and '"""
            + rdate2
            + """'
            AND fcode like  '"""
            + fcode
            + """%'
            AND fstatus like  '"""
            + fstatus
            + """%'
            
            union all
            
            SELECT a.fname,   
                    @row:=@row+1 as No,
                " " as rcheck,
                "DB" as fpath,
                    a.rdate,   
                    a.fcode,   
                    a.fstatus,   
                    a.in_date  
            FROM kradata a,(select @row :=0 from dual) b
            WHERE Left( Right( fname , 6), 2) like '"""
            + rcity
            + """%'
            AND rdate between '"""
            + rdate1
            + """' and '"""
            + rdate2
            + """'
            AND fcode like  '"""
            + fcode
            + """%'
            AND fstatus like  '"""
            + fstatus
            + """%'
            ; """
        )

        r_cnt = cursor.execute(strSql)         # 결과값 개수 반환
        result = cursor.fetchall()

        # connection.commit()
        # connection.close()

        # print(strSql)
        # print(result[0:5])

    except:
        # connection.rollback()
        print("Failed selecting in krafile")
    finally:
        if cursor:
            cursor.close()

    return result


def get_kradata(rcity, rdate1, rdate2, fcode, fstatus):

    try:
        cursor = connection.cursor()

        strSql = """ 
            SELECT a.fname,   
                    @row:=@row+1 as No,
                " " as rcheck,
                "" as fpath,
                    a.rdate,   
                    a.fcode,   
                    a.fstatus,   
                    a.in_date  
            FROM kradata a,(select @row :=0 from dual) b
            WHERE Left( Right( fname , 6), 2) like '""" + rcity + """%'
            AND rdate between '""" + rdate1 + """' and '""" + rdate2 + """'
            AND fcode like  '""" + fcode + """%'
            AND fstatus like  '""" + fstatus + """%'
            ; """

        r_cnt = cursor.execute(strSql)         # 결과값 개수 반환
        result = cursor.fetchall()

        # connection.commit()
        # connection.close()

        # print(strSql)
        # print(result[0:5])

    except:
        # connection.rollback()
        print("Failed selecting in krafile")
    finally:
        if cursor:
            cursor.close()

    return result


def get_breakingnews(rcity, rdate1, rdate2, title):

    try:
        cursor = connection.cursor()

        strSql = """ 
              SELECT @row:=@row+1 as No,
                  " " as rcheck,
                      a.rcity,   
                      a.rdate,   
                      a.title,
                      a.news,   
                      a.in_date  
              FROM breakingnews a,(select @row :=0 from dual) b
              WHERE rcity like '""" + rcity + """%'
              AND rdate between '""" + rdate1 + """' and '""" + rdate2 + """'
              AND title like  '""" + title + """%'
              
            ; """

        print(strSql)
        r_cnt = cursor.execute(strSql)         # 결과값 개수 반환
        result = cursor.fetchall()

        connection.commit()
        connection.close()

    except:
        connection.rollback()
        print("Failed selecting in Breaking News")

    return result


def get_file_contents(fname):

    try:
        cursor = connection.cursor()

        strSql = """ 
              SELECT fcontents
              FROM kradata  
              WHERE fname = '""" + fname + """'
            ; """

        r_cnt = cursor.execute(strSql)         # 결과값 개수 반환
        result = cursor.fetchone()
        # result = cursor.fetchall()

        connection.commit()
        connection.close()

        # print(strSql)
        # print(result[0:5])

    except:
        connection.rollback()
        print("Failed selecting in krafile")

    return str(result[0], 'cp949')


def krafile_convert(fnames):

    for fname in fnames:

        if fname[-4:] == 'xlsx':

            convert_c1(fname)

        else:
            if fname[-12:-10] == '11':
                print(fname[-12:-10])

                file = open(fname, "r")
                while True:
                    line = file.readline()
                    if not line:
                        break
                    print(line.strip())
                file.close()

    return (fnames)


def convert_c1(fname):

    print(fname)
    # df = pd.read_excel(fname,  index_col=None, engine='openpyxl')
    wb = op.load_workbook(fname)
    ws = wb.active

    # print(ws.max_row)
    # print(ws.max_column)
    rcity = fname[-10:-5]

    tdate = fname[-19:-11]

    # print(tdate)

    try:
        cursor = connection.cursor()

        strSql = """
                DELETE FROM train
                WHERE rcity = '""" + rcity + """' and tdate = '""" + tdate + """'
            ; """

        # print(strSql)

        r_cnt = cursor.execute(strSql)         # 결과값 개수 반환
        result = cursor.fetchone()
        # result = cursor.fetchall()
        # print(r_cnt)

        # connection.commit()
        # connection.close()

    except:
        # connection.rollback()
        print("Failed Deleting in train Table")

    cnt = 0
    for row_rng in ws.rows:

        team = row_rng[1].value
        team_num = row_rng[2].value
        horse = row_rng[3].value
        grade = row_rng[4].value
        rider = row_rng[5].value
        in_time = row_rng[6].value
        out_time = row_rng[7].value
        t_time = row_rng[8].value[0:-1]

        canter = row_rng[9].value[2:3]
        strong = row_rng[9].value[6:7]

        remark = row_rng[10].value

        try:
            cursor = connection.cursor()

            strSql = """
                    INSERT INTO train
                    ( rcity, tdate, horse, team, team_num, grade, rider, in_time, out_time, t_time, canter, strong, remark )
                    VALUES
                    ( '""" + rcity + """',
                        '""" + tdate + """',
                        '""" + horse + """',
                        '""" + team + """',
                        '""" + team_num + """',
                        '""" + grade + """',
                        '""" + rider + """',
                        '""" + in_time + """',
                        '""" + out_time + """',
                        """ + t_time + """,
                        """ + canter + """,
                        """ + strong + """,
                        '""" + remark + """'
                    ) ; """

            # print(strSql)

            r_cnt = cursor.execute(strSql)         # 결과값 개수 반환
            result = cursor.fetchone()
            # result = cursor.fetchall()

            # connection.commit()
            # connection.close()

            cnt += 1

        except:
            # connection.rollback()
            print("Failed inserting in train Table")

    #     for cell in row_rng:  # 각 행에 대한 1차원 배열 for문
    #         print(cell.value)
    return (cnt)
